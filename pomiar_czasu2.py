#sciągnij chromedrivera, wrzuc go folderu z localem chrome
#https://www.youtube.com/watch?v=FVumnHy5Tzo

from selenium import webdriver
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import zipfile

cwd = r'C:\Program Files\Google\Chrome\Application'
os.chdir(cwd)
os.system('start /B start cmd.exe @cmd /k chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\API\localhost"')

option = webdriver.ChromeOptions()
option.add_experimental_option("debuggerAddress", "localhost:9222")
try:
    driver = webdriver.Chrome(executable_path=r'C:\API\chromedriver.exe', options = option) 
except:
    url = 'https://chromedriver.chromium.org/'
    reqs = requests.get(url)
    soup = BeautifulSoup(reqs.text, 'html.parser')    
    urls = []
    n = 0
    for link in soup.find_all('a'):
        a = link.get('href')
        try:
            if 'https://chromedriver.storage.googleapis.com/index.html?path=' in a:
                n = n + 1
                if n == 2:
                    wersja = link.get('href')
        except TypeError:
            continue
        
    pobierz = 'https://chromedriver.storage.googleapis.com/' + wersja[60:] + 'chromedriver_win32.zip'
    
    r = requests.get(pobierz, allow_redirects=True)
    
    open(r'C:\API\chromium.zip', 'wb').write(r.content)
    
    path_to_zip_file = r'C:\API\chromium.zip'
    directory_to_extract_to = r'C:\API'
    
    with zipfile.ZipFile(path_to_zip_file, 'r') as zip_ref:
        zip_ref.extractall(directory_to_extract_to)

driver = webdriver.Chrome(executable_path=r'C:\API\chromedriver.exe', options = option)   
driver.current_url
driver.get("https://www.google.com/intl/pl/gmail/about/#")
flaga = False
czas_poczatku = datetime.now()

while not flaga:
    try:
        a =  driver.title
    except:
        flaga = True
        czas_konca = datetime.now()
roznica = czas_konca - czas_poczatku

d = {'czas_poczatku':[czas_poczatku], \
     'czas_konca':[czas_konca], \
     'roznica':str(czas_konca - czas_poczatku)}
   
df = pd.DataFrame(data = d)

df['roznica'] = pd.to_datetime(df['roznica']).dt.strftime("%H:%M:%S")

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):

    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

append_df_to_excel(r'C:\MyPythonScripts\pomiar_czasu\pomiar_czasu.xlsx', df, header=None, index=False)

